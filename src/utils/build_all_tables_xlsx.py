# -*- coding: utf-8 -*-
"""All manuscript tables into ONE Excel workbook, one sheet each, formatted like the
merged-cohort file (dark header, freeze header, autofilter, borders, column widths)."""
import os
from math import floor, log10

import pandas as pd
from scipy import stats
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
RES = os.path.join(BASE, "data", "04_results")
OUT = os.path.join(BASE, "paper", "UTI_main_tables.xlsx")
TARGET = "CULTIVO_PATOLOGICO"


def sig(x, n=3):
    if pd.isna(x):
        return ""
    if x == 0:
        return "0"
    d = n - 1 - floor(log10(abs(x)))
    return f"{round(x, d):.0f}" if d <= 0 else f"{x:.{d}f}"


# ---------------------------------------------------------------------------
# Table 1: cohort characteristics (recomputed, 3 sig figs, * = Boruta-retained)
# ---------------------------------------------------------------------------
RETAINED = {"Age, years", "Sex", "Leukocyte esterase", "Nitrite", "Protein",
            "Blood (Hb peroxidase)", "Specific gravity", "Bacteria, /µL",
            "White blood cells, /µL", "Red blood cells, /µL", "Epithelial cells, /µL",
            "Sysmex Gram flag"}
raw = pd.read_excel(os.path.join(BASE, "data", "01_raw", "uti_raw.xlsx")); raw["RAW_INDEX"] = raw.index + 1
ml = pd.read_excel(os.path.join(BASE, "data", "03_processed", "uti_ml_final.xlsx"))
keep = ["EDAD", "SEXO", "LEUT", "NITT", "PROTT", "HEMATT", "GLUT", "PHT", "DENST",
        "RBO", "WBCO", "EC", "BACTS", "CASTS", "YLC", "BACT_INFO"]
c = ml[["RAW_INDEX", TARGET]].merge(raw[["RAW_INDEX"] + keep], on="RAW_INDEX", how="left")
c["SEXO"] = c["SEXO"].map({"H": "Male", "M": "Female"})
for col in ["LEUT", "NITT", "PROTT", "HEMATT", "GLUT"]:
    c[col] = c[col].astype(str).replace({"Negativo": "Negative", "Positivo": "Positive"})
def gram(x):
    s = str(x).lower()
    if "sin informaci" in s: return "No information"
    if "negativo" in s and ("positivo" in s or "gran positivo" in s): return "Mixed"
    if "positivo" in s: return "Gram-positive"
    if "negativo" in s: return "Gram-negative"
    return "No information"
c["GRAM"] = c["BACT_INFO"].map(gram)
neg, pos = c[c[TARGET] == 0], c[c[TARGET] == 1]
T1COLS = ["Characteristic", f"Culture-negative (n={len(neg)})",
          f"Culture-positive (n={len(pos)})", "p-value"]
t1 = []
def star(l): return l + " *" if l in RETAINED else l
def cont(label, col):
    p = stats.mannwhitneyu(neg[col].dropna(), pos[col].dropna()).pvalue
    def mi(s): return f"{sig(s.median())} ({sig(s.quantile(.25))}-{sig(s.quantile(.75))})"
    t1.append([star(label), mi(neg[col]), mi(pos[col]), "<0.05" if p < 0.05 else f"{p:.2f}"])
def cat(label, col, order):
    p = stats.chi2_contingency(pd.crosstab(c[col], c[TARGET]).reindex(order).fillna(0).values)[1]
    t1.append([star(label), "", "", "<0.05" if p < 0.05 else f"{p:.2f}"])
    for k in order:
        nn, pp = (neg[col] == k).sum(), (pos[col] == k).sum()
        t1.append([f" {k}", f"{nn} ({nn/len(neg)*100:.1f}%)",
                   f"{pp} ({pp/len(pos)*100:.1f}%)", ""])
cont("Age, years", "EDAD"); cat("Sex", "SEXO", ["Male", "Female"])
cat("Leukocyte esterase", "LEUT", ["Negative", "25", "75", "500"])
cat("Nitrite", "NITT", ["Negative", "Positive"])
cat("Protein", "PROTT", ["Negative", "15", "30", "100", "300", "1000"])
cat("Blood (Hb peroxidase)", "HEMATT", ["Negative", "10", "20", "50", "250"])
cat("Glucose", "GLUT", ["Negative", "Positive"]); cont("pH", "PHT"); cont("Specific gravity", "DENST")
cont("Bacteria, /µL", "BACTS"); cont("White blood cells, /µL", "WBCO"); cont("Red blood cells, /µL", "RBO")
cont("Epithelial cells, /µL", "EC"); cont("Hyaline casts, /µL", "CASTS"); cont("Yeasts, /µL", "YLC")
cat("Sysmex Gram flag", "GRAM", ["Gram-negative", "Gram-positive", "Mixed", "No information"])
table1 = pd.DataFrame(t1, columns=T1COLS)

# ---------------------------------------------------------------------------
# Table 2: performance across the two phases
# ---------------------------------------------------------------------------
ci = pd.read_excel(os.path.join(RES, "ci_and_fairness.xlsx"), sheet_name="ci_overall")
w1 = pd.read_excel(os.path.join(RES, "phase2_wave1.xlsx"), sheet_name="performance")
# bootstrap 95% CIs for the Phase-1 matched-subset rows; see src/utils/compute_table2_ci.py
t2ci = pd.read_excel(os.path.join(RES, "table2_ci.xlsx"), keep_default_na=False).set_index("Row")
# post-implementation row (n=197, Option A) from the 2026-07-23 prospective analysis (uti_23)
pp = pd.read_excel(os.path.join(RES, "prospective_2026-07-23.xlsx"),
                   sheet_name="table2_row", keep_default_na=False).iloc[0]
def gc(row, col): return t2ci.loc[row, col]
def g(df, col, r=0): return df.iloc[r][col] if col in df.columns else ""
R_CONC = "Phase 1: model, concordant subset"
R_PREV = "Phase 1: previous model (combined decision)"
table2 = pd.DataFrame([
    [1, "Hold-out (development)", "Retrospective", int(ci.iloc[0]["N"]), g(ci, "ROC AUC"), g(ci, "Sensitivity"), g(ci, "Specificity"), g(ci, "MCC")],
    [1, "New model (matched subset) *", "Retrospective", int(t2ci.loc[R_CONC, "N"]), gc(R_CONC, "ROC AUC"), gc(R_CONC, "Sensitivity"), gc(R_CONC, "Specificity"), gc(R_CONC, "MCC")],
    [1, "Previous model (matched subset) *", "Retrospective", int(t2ci.loc[R_PREV, "N"]), gc(R_PREV, "ROC AUC"), gc(R_PREV, "Sensitivity"), gc(R_PREV, "Specificity"), gc(R_PREV, "MCC")],
    [2, "Pre-implementation batch", "Prospective (5-15 Jun 2026)", int(w1.iloc[0]["N"]), g(w1, "ROC AUC"), g(w1, "Sensitivity"), g(w1, "Specificity"), g(w1, "MCC")],
    [2, "Post-implementation (routine use)", "Prospective (22 Jun-22 Jul 2026)", int(pp["N"]), pp["ROC AUC"], pp["Sensitivity"], pp["Specificity"], pp["MCC"]],
], columns=["Phase", "Dataset", "Population", "n", "ROC AUC", "Sensitivity", "Specificity", "MCC"])
# uniform CI format: parentheses (not square brackets), and point estimate over CI on a second line in the SAME cell
for mc_ in ["ROC AUC", "Sensitivity", "Specificity", "MCC"]:
    table2[mc_] = (table2[mc_].astype(str)
                   .str.replace("[", "(", regex=False).str.replace("]", ")", regex=False)
                   .str.replace(" (", "\n(", regex=False))

# ---------------------------------------------------------------------------
# Supplementary tables (read as-is)
# ---------------------------------------------------------------------------
def rd(f, sheet=0): return pd.read_excel(os.path.join(RES, f), sheet_name=sheet)
SHEETS = [
    ("Table 1. Cohort", table1),
    ("Table 2. Performance", table2),
]

# ---------------------------------------------------------------------------
# Write + format
# ---------------------------------------------------------------------------
hdr_fill = PatternFill("solid", fgColor="1E3A5F")
hdr_font = Font(color="FFFFFF", bold=True, size=11)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="D0D7DE")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

with pd.ExcelWriter(OUT, engine="openpyxl") as xw:
    for name, df in SHEETS:
        df.to_excel(xw, sheet_name=name[:31], index=False)
        ws = xw.sheets[name[:31]]
        nrow, ncol = df.shape
        for j, col in enumerate(df.columns, 1):
            cell = ws.cell(1, j); cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = center; cell.border = border
        for i in range(2, nrow + 2):
            for j in range(1, ncol + 1):
                cc = ws.cell(i, j); cc.border = border
                cc.alignment = left if j == 1 else center
        # column widths from the widest single line (cells may contain '\n')
        for j, col in enumerate(df.columns, 1):
            longest = max([len(str(col))] + [len(ln) for v in df.iloc[:, j - 1]
                                             for ln in str(v).split("\n")])
            ws.column_dimensions[get_column_letter(j)].width = min(longest + 2, 46)
        ws.row_dimensions[1].height = 28
        # taller data rows so two-line cells (value over CI) fit
        multiline = df.astype(str).apply(lambda s: s.str.contains("\n")).any().any()
        if multiline:
            for i in range(2, nrow + 2):
                ws.row_dimensions[i].height = 34
        # Table 2: merge the Phase column so each phase number spans its rows
        if df.columns[0] == "Phase":
            phases = list(df["Phase"])
            i = 0
            while i < len(phases):
                j2 = i
                while j2 + 1 < len(phases) and phases[j2 + 1] == phases[i]:
                    j2 += 1
                if j2 > i:  # run of >1 row -> merge (Excel rows are df-index + 2)
                    ws.merge_cells(f"A{i + 2}:A{j2 + 2}")
                i = j2 + 1
            ws.column_dimensions["A"].width = 7
            # footnote beneath the table
            fr = nrow + 3
            ws.merge_cells(start_row=fr, start_column=1, end_row=fr, end_column=ncol)
            fc = ws.cell(fr, 1)
            fc.value = (
                "* Matched subset: the 1,248 hold-out samples for which the previous model issued "
                "a decision; the new-model and previous-model rows are computed on the same "
                "patients. Previous model = PRED_CULT_IA, the combined decision of UTI_CDS and "
                "CDS_RNA; NA = ROC AUC is not defined for its binary decision. Each cell shows the "
                "point estimate with its 95% bootstrap confidence interval below. Post-implementation "
                "performance uses the live deployed probability against urine culture "
                "(CULTIVO_PATOLOGICO, the same reference as in development).")
            fc.font = Font(italic=True, size=9)
            fc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[fr].height = 58
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{nrow + 1}"

print("Sheets:", [s[0] for s in SHEETS])
print("Saved:", OUT)
