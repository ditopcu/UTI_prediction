# -*- coding: utf-8 -*-
"""
Merge ALL patients (matched + unmatched) from the two post-implementation files
into a single, cleanly formatted Excel workbook.

Outer join on the lab petition number so every patient appears once, tagged with
its match status. Junk / empty / admin columns are dropped.
"""
import os

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(BASE, "data", "05_prospective", "2026-06_postimpl")
CB_FILE = os.path.join(DATA, "catboost_live_predictions.xlsx")
DE_FILE = os.path.join(DATA, "culture_results.xlsx")
OUT = os.path.join(DATA, "all_patients_merged.xlsx")

# ---------------------------------------------------------------------------
# Load
# ---------------------------------------------------------------------------
cb = pd.read_excel(CB_FILE)
de = pd.read_excel(DE_FILE)

prob_col = [c for c in cb.columns if c.startswith("Probabilidad")][0]
pred_col = [c for c in cb.columns if c.startswith("Predicci") and "Resultado" not in c][0]
pet_col = [c for c in cb.columns if "eticion" in c][0]

cb_slim = cb[[pet_col, "Fecha", "Paciente", prob_col, pred_col]].rename(
    columns={pet_col: "PETICION", "Fecha": "FECHA_CB",
             "Paciente": "PACIENTE_CB", prob_col: "CATBOOST_PROB", pred_col: "CATBOOST_PRED"}
)

de_keep = de.rename(columns={"PETICIONCB": "PETICION"})[[
    "PETICION", "FECHA", "PACIENTE", "EDAD", "SEXO", "SERVICIO", "PROCEDENCIA",
    "CULTIVO_PATOLOGICO", "RESULTADO_CULTIVO", "FECHA_CULTIVO",
    "UTI_CDS", "CDS_RNA", "PRED_CULT_IA",
    "LEUT", "NITT", "DENST", "PHT", "PROTT", "GLUT", "HEMATT", "RBO", "WBCO",
    "EC", "BACTS", "CASTS", "YLC", "UTI_INFO", "BACT_INFO", "LEU", "NEUT", "PCR",
]].copy()

# Simplify the free-text hospital-algorithm outputs to short Alta/Baja labels
def short_label(x):
    if pd.isna(x):
        return np.nan
    t = str(x).lower()
    if "alta" in t:
        return "Alta"
    if "baja" in t:
        return "Baja"
    return np.nan  # drops the one malformed CDS_RNA cell

de_keep["CDS_RNA"] = de_keep["CDS_RNA"].map(short_label)
de_keep["PRED_CULT_IA"] = de_keep["PRED_CULT_IA"].map(short_label)

# ---------------------------------------------------------------------------
# Outer join + coalesce identity fields
# ---------------------------------------------------------------------------
m = de_keep.merge(cb_slim, on="PETICION", how="outer")
m["FECHA"] = m["FECHA"].fillna(m["FECHA_CB"])
m["PACIENTE"] = m["PACIENTE"].fillna(m["PACIENTE_CB"])
m = m.drop(columns=["FECHA_CB", "PACIENTE_CB"])

in_cb = m["CATBOOST_PROB"].notna()
in_de = m["CULTIVO_PATOLOGICO"].notna()

m["ESTADO"] = np.select(
    [in_cb & in_de, in_cb & ~in_de, ~in_cb & in_de],
    ["Eslesti", "Solo CatBoost", "Solo Cultivo"], default="?",
)

# TP/TN/FP/FN + correctness for matched rows
def classify(r):
    if not (pd.notna(r["CATBOOST_PRED"]) and pd.notna(r["CULTIVO_PATOLOGICO"])):
        return ""
    y, yhat = int(r["CULTIVO_PATOLOGICO"]), int(r["CATBOOST_PRED"])
    return {(1, 1): "TP", (0, 0): "TN", (0, 1): "FP", (1, 0): "FN"}[(y, yhat)]

m["CLASIFICACION"] = m.apply(classify, axis=1)
m["CORRECTO"] = np.where(m["CLASIFICACION"].isin(["TP", "TN"]), "Si",
                np.where(m["CLASIFICACION"].isin(["FP", "FN"]), "No", ""))

# ---------------------------------------------------------------------------
# Column order + sort
# ---------------------------------------------------------------------------
ORDER = [
    "ESTADO", "PETICION", "FECHA", "PACIENTE", "EDAD", "SEXO", "SERVICIO", "PROCEDENCIA",
    "CATBOOST_PROB", "CATBOOST_PRED", "CULTIVO_PATOLOGICO", "CLASIFICACION", "CORRECTO",
    "UTI_CDS", "CDS_RNA", "PRED_CULT_IA",
    "RESULTADO_CULTIVO", "FECHA_CULTIVO",
    "LEUT", "NITT", "DENST", "PHT", "PROTT", "GLUT", "HEMATT", "RBO", "WBCO",
    "EC", "BACTS", "CASTS", "YLC", "UTI_INFO", "BACT_INFO", "LEU", "NEUT", "PCR",
]
m = m[ORDER]
estado_rank = {"Eslesti": 0, "Solo Cultivo": 1, "Solo CatBoost": 2}
m = m.sort_values(by=["ESTADO", "FECHA", "PETICION"],
                  key=lambda s: s.map(estado_rank) if s.name == "ESTADO" else s
                  ).reset_index(drop=True)

# ---------------------------------------------------------------------------
# Write + format
# ---------------------------------------------------------------------------
with pd.ExcelWriter(OUT, engine="openpyxl") as xw:
    m.to_excel(xw, index=False, sheet_name="Todos")
    ws = xw.sheets["Todos"]

    nrow, ncol = m.shape
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    for j, col in enumerate(m.columns, start=1):
        c = ws.cell(row=1, column=j)
        c.fill, c.font, c.alignment, c.border = header_fill, header_font, center, border

    # Status / classification colors
    estado_fill = {
        "Eslesti": PatternFill("solid", fgColor="E7F0FF"),
        "Solo Cultivo": PatternFill("solid", fgColor="FFF4E5"),
        "Solo CatBoost": PatternFill("solid", fgColor="F0F0F0"),
    }
    clasif_fill = {
        "TP": PatternFill("solid", fgColor="D6F5D6"),
        "TN": PatternFill("solid", fgColor="D6F5D6"),
        "FP": PatternFill("solid", fgColor="FBD5D5"),
        "FN": PatternFill("solid", fgColor="FBD5D5"),
    }
    col_idx = {col: i + 1 for i, col in enumerate(m.columns)}

    for i in range(nrow):
        excel_row = i + 2
        est = m.iloc[i]["ESTADO"]
        cla = m.iloc[i]["CLASIFICACION"]
        # tint the ESTADO cell
        ws.cell(row=excel_row, column=col_idx["ESTADO"]).fill = estado_fill.get(est, PatternFill())
        # tint the CLASIFICACION + CORRECTO cells
        if cla in clasif_fill:
            ws.cell(row=excel_row, column=col_idx["CLASIFICACION"]).fill = clasif_fill[cla]
            ws.cell(row=excel_row, column=col_idx["CORRECTO"]).fill = clasif_fill[cla]
        for j in range(1, ncol + 1):
            ws.cell(row=excel_row, column=j).border = border

    # Number formats + alignment
    prob_letter = get_column_letter(col_idx["CATBOOST_PROB"])
    for r in range(2, nrow + 2):
        ws[f"{prob_letter}{r}"].number_format = "0.00"
    for name in ["ESTADO", "PETICION", "EDAD", "SEXO", "CATBOOST_PROB", "CATBOOST_PRED",
                 "CULTIVO_PATOLOGICO", "CLASIFICACION", "CORRECTO",
                 "UTI_CDS", "CDS_RNA", "PRED_CULT_IA"]:
        L = get_column_letter(col_idx[name])
        for r in range(2, nrow + 2):
            ws[f"{L}{r}"].alignment = center

    # Column widths
    widths = {
        "ESTADO": 14, "PETICION": 11, "FECHA": 12, "PACIENTE": 34, "EDAD": 6, "SEXO": 6,
        "SERVICIO": 22, "PROCEDENCIA": 26, "CATBOOST_PROB": 9, "CATBOOST_PRED": 8,
        "CULTIVO_PATOLOGICO": 10, "CLASIFICACION": 8, "CORRECTO": 9,
        "UTI_CDS": 9, "CDS_RNA": 9, "PRED_CULT_IA": 12,
        "RESULTADO_CULTIVO": 40, "FECHA_CULTIVO": 13, "BACT_INFO": 24,
    }
    for col in m.columns:
        L = get_column_letter(col_idx[col])
        ws.column_dimensions[L].width = widths.get(col, 9)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "D2"                       # freeze header + first 3 id cols
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{nrow + 1}"

# ---------------------------------------------------------------------------
# Console summary
# ---------------------------------------------------------------------------
counts = m["ESTADO"].value_counts()
print("Total pacientes:", nrow)
for k in ["Eslesti", "Solo Cultivo", "Solo CatBoost"]:
    print(f"  {k:14s}: {counts.get(k, 0)}")
mm = m[m["ESTADO"] == "Eslesti"]
print("Matched CLASIFICACION:", mm["CLASIFICACION"].value_counts().to_dict())
print("Saved:", OUT)
